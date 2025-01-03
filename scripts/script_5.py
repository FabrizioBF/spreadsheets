# Ler células de uma linha específica
from openpyxl import load_workbook


def iterating_row(path, sheet_name, row):
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f"'{sheet_name}' não encontrado. Encerrando.")
        return
    sheet = workbook[sheet_name]
    for cell in sheet[row]:
        print(f"{cell.column_letter}{cell.row} = {cell.value}")


if __name__ == "__main__":
    iterating_row("spreadsheets/planilhas/livros.xlsx",
                  sheet_name="mercadorias", row=2)
