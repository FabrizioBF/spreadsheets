# Ler células de um intervalo
import openpyxl
from openpyxl import load_workbook


def iterating_over_values(path, sheet_name, cell_range):
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f"'{sheet_name}' não encontrado. Encerrando.")
        return
    sheet = workbook[sheet_name]
    for column in sheet[cell_range]:
        for cell in column:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            print(f"{cell.column_letter}{cell.row} = {cell.value}")


if __name__ == "__main__":
    iterating_over_values("spreadsheets/planilhas/livros.xlsx", sheet_name="mercadorias",
                          cell_range="A1:B6")
