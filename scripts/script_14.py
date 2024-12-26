# Criar os títulos das planilhas
from openpyxl import Workbook


def create_sheets(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Finanças"
    sheet2 = workbook.create_sheet(title="Inovação")
    workbook.save(path)


if __name__ == "__main__":
    create_sheets("departamentos.xlsx")
    print("Arquivo criado com sucesso!")
