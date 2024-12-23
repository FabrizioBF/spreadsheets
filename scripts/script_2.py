#Tarefa: abrir uma planilha;
from openpyxl import load_workbook


print("Abrindo um planilha")

def open_workbook(path):
    workbook = load_workbook(filename=path)
    print(f"Nomes das planilhas (abas): {workbook.sheetnames}")
    sheet = workbook.active
    print(f"O título da primeira planilha (aba) é: {sheet.title}")
    
    
if __name__ == "__main__":
    open_workbook("spreadsheets/planilhas/livros.xlsx")