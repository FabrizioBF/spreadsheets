# Ler células de várias linhas ou colunas
from openpyxl import load_workbook


def iterating_over_values(path, sheet_name):
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f"'{sheet_name}' não encontrado. Encerrando.")
        return
    sheet = workbook[sheet_name]
    print("\nLer células de várias linhas ou colunas")
    for value in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3, values_only=True):
        print(value)


if __name__ == "__main__":
    iterating_over_values(
        "spreadsheets/planilhas/livros.xlsx", sheet_name="vendas")
