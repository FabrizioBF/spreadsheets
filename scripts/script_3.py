# Tafera: ler uma planilha específica
from openpyxl import load_workbook


def open_workbook(path, sheet_name):
    workbook = load_workbook(filename=path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"\nO título da planilha é: {sheet.title}")
        print(f"As células que contêm dados: {sheet.calculate_dimension()}")


if __name__ == "__main__":
    open_workbook("spreadsheets/planilhas/livros.xlsx", sheet_name="vendas")
