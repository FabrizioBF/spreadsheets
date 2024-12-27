import openpyxl


def create_worksheets(path):
    workbook = openpyxl.Workbook()
    workbook.create_sheet(title="Primeira planilha")
    print(workbook.sheetnames)
    # Insert a worksheet
    workbook.create_sheet(index=1, title="Segunda planilha")
    print(workbook.sheetnames)
    del workbook["Segunda planilha"]
    print(workbook.sheetnames)
    workbook.save(path)


if __name__ == "__main__":
    create_worksheets("planilha_trabalho.xlsx")
